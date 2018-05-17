import time
import datetime
import json
import os
import re
import sys
import errno
import uuid
import zipfile
from Bio import SeqIO
from six import string_types
import xlsxwriter
from openpyxl import load_workbook

from DataFileUtil.DataFileUtilClient import DataFileUtil
from AssemblyUtil.AssemblyUtilClient import AssemblyUtil
from KBaseReport.KBaseReportClient import KBaseReport
from SetAPI.SetAPIClient import SetAPI
from biokbase.workspace.client import Workspace as workspaceService
from pprint import pformat


def log(message, prefix_newline=False):
    """Logging function, provides a hook to suppress or redirect log messages."""
    print(('\n' if prefix_newline else '') + '{0:.2f}'.format(time.time()) + ': ' + str(message))


class MetagenomeFileUtils:

    def _validate_merge_bins_from_binned_contig_params(self, params):
        """
        _validate_merge_bins_from_binned_contig_params:
                validates params passed to merge_bins_from_binned_contig method

        """
        log('Start validating merge_bins_from_binned_contig params')

        # check for required parameters
        for p in ['old_binned_contig_ref', 'bin_merges',
                  'output_binned_contig_name', 'workspace_name']:
            if p not in params:
                raise ValueError('"{}" parameter is required, but missing'.format(p))

        bin_merges = params.get('bin_merges')

        if not isinstance(bin_merges, list):
            error_msg = 'expecting a list for bin_merges param, '
            error_msg += 'but getting a [{}]'.format(type(bin_merges))
            raise ValueError(error_msg)

        for bin_merge in bin_merges:
            for p in ['new_bin_id', 'bin_to_merge']:
                if p not in bin_merge:
                    raise ValueError('"{}" key is required in bin_merges, but missing'.format(p))

            bin_to_merge = bin_merge.get('bin_to_merge')

            if not isinstance(bin_to_merge, list):
                error_msg = 'expecting a list for bin_to_merge, '
                error_msg += 'but getting a [{}]'.format(type(bin_to_merge))
                raise ValueError(error_msg)

    def _validate_remove_bins_from_binned_contig_params(self, params):
        """
        _validate_remove_bins_from_binned_contig_params:
                validates params passed to remove_bins_from_binned_contig method

        """
        log('Start validating remove_bins_from_binned_contig params')

        # check for required parameters
        for p in ['old_binned_contig_ref', 'bins_to_remove',
                  'output_binned_contig_name', 'workspace_name']:
            if p not in params:
                raise ValueError('"{}" parameter is required, but missing'.format(p))

        bins_to_remove = params.get('bins_to_remove')

        if not isinstance(bins_to_remove, list):
            error_msg = 'expecting a list for bins_to_remove param, '
            error_msg += 'but getting a [{}]'.format(type(bins_to_remove))
            raise ValueError(error_msg)

    def _validate_file_to_binned_contigs_params(self, params):
        """
        _validate_file_to_binned_contigs_params:
                validates params passed to file_to_binned_contigs method

        """
        log('Start validating file_to_binned_contigs params')

        # check for required parameters
        for p in ['assembly_ref', 'file_directory', 'binned_contig_name', 'workspace_name']:
            if p not in params:
                raise ValueError('"{}" parameter is required, but missing'.format(p))

    def _validate_binned_contigs_to_file_params(self, params):
        """
        _validate_binned_contigs_to_file_params:
                validates params passed to binned_contigs_to_file method

        """

        log('Start validating binned_contigs_to_file params')

        # check for required parameters
        for p in ['input_ref']:
            if p not in params:
                raise ValueError('"{}" parameter is required, but missing'.format(p))

    def _validate_import_excel_as_binned_contigs_params(self, params):
        """
        _validate_import_excel_as_binned_contigs_params:
                validates params passed to import_excel_as_binned_contigs method

        """

        log('Start validating import_excel_as_binned_contigs params')

        # check for required parameters
        for p in ['shock_id', 'workspace_name']:
            if p not in params:
                raise ValueError('"{}" parameter is required, but missing'.format(p))

    def _validate_extract_binned_contigs_as_assembly_params(self, params):
        """
        _validate_extract_binned_contigs_as_assembly_params:
                validates params passed to extract_binned_contigs_as_assembly method

        """

        log('Start validating extract_binned_contigs_as_assembly params')

        # check for required parameters
        for p in ['binned_contig_obj_ref', 'extracted_assemblies',
                  'assembly_suffix', 'workspace_name']:
            if p not in params:
                raise ValueError('"{}" parameter is required, but missing'.format(p))

        # convert comma-separated list of bins into a list of individual ids (the python
        # comprehension construction deals with the fact that split(',') returns a list of
        # length one, [''], for an empty string input

        extracted_assemblies = [x for x in params.get('extracted_assemblies').split(',') if x]

        # parameter assembly_set_name is required if extracted_assemblies list has more
        # than one element

        if len(extracted_assemblies) > 1 and 'assembly_set_name' not in params:
            raise ValueError(
                '"assembly_set_names" parameter is required for more than one extracted assembly')

    def _mkdir_p(self, path):
        """
        _mkdir_p: make directory for given path
        """
        if not path:
            return
        try:
            os.makedirs(path)
        except OSError as exc:
            if exc.errno == errno.EEXIST and os.path.isdir(path):
                pass
            else:
                raise

    def _get_bin_ids(self, file_directory):
        """
        _get_bin_ids: getting bin contig ids from files

        NOTE: This method is very specific to MaxBin2 app result.
              Bin contig files generated by MaxBin2 follow 'header.0xx.fasta' name pattern
        """

        bin_ids = []

        result_files = os.listdir(file_directory)

        for file in result_files:
            if re.match(r'.*\.\d{3}\.fasta', file):
                bin_ids.append(file)

        log('generated bin ids:\n{}'.format('\n'.join(bin_ids)))

        return bin_ids

    def _process_summary_file(self, bin_id, lines):
        """
        _process_summary_file: process header.summary file content
                               getting GC content (gc), Genome size (sum_contig_len)
                               and Completeness (cov) from header.summary file

        NOTE: This method is very specific to MaxBin2 app result.

        header.summary file could be one of below fomat:
        Bin name                  Abundance  Completeness    Genome size     GC content
        maxbin_output.001.fasta   0.00       97.2%           2690533         52.9

        Bin name                  Completeness    Genome size     GC content
        maxbin_output.001.fasta   97.2%           2690533         52.9
        """

        for line in lines:
            line_list = line.split('\t')
            if line_list[0] == bin_id:
                if len(line_list) == 5:
                    gc = round(float(line_list[4]) / 100, 5)
                    sum_contig_len = int(line_list[3])
                    cov = round(float(line_list[2].partition('%')[0]) / 100, 5)
                elif len(line_list) == 4:
                    gc = round(float(line_list[3]) / 100, 5)
                    sum_contig_len = int(line_list[2])
                    cov = round(float(line_list[1].partition('%')[0]) / 100, 5)

        return gc, sum_contig_len, cov

    def _get_total_contig_len(self, file_directory):
        """
        _get_total_contig_len: process header.summary file content
                               getting total contig length from header.summary file

        NOTE: This method is very specific to MaxBin2 app result.
        """

        log('generating total contig length')
        total_contig_len = 0

        file_list = os.listdir(file_directory)
        for file in file_list:
            if file.endswith('.summary'):
                with open(os.path.join(file_directory, file), 'r') as summary_file:
                    lines = summary_file.readlines()
                    for line in lines[1:]:
                        line_list = line.split('\t')
                        if len(line_list) == 5:
                            total_contig_len += int(line_list[3])
                        elif len(line_list) == 4:
                            total_contig_len += int(line_list[2])

        log('generated total contig length: {}'.format(total_contig_len))
        return total_contig_len

    def _generate_contig_bin_summary(self, bin_id, file_directory):
        """
        _generate_contig_bin_summary: getting ContigBin summary from header.summary file

        NOTE: This method is very specific to MaxBin2 app result.
        """
        log('generating summary for bin_id: {}'.format(bin_id))

        file_list = os.listdir(file_directory)

        for file in file_list:
            if file.endswith('.summary'):
                with open(os.path.join(file_directory, file), 'r') as summary_file:
                    lines = summary_file.readlines()
                    gc, sum_contig_len, cov = self._process_summary_file(bin_id, lines)

        log('generated GC content: {}, Genome size: {} '.format(gc, sum_contig_len))
        log('and Completeness: {} for bin_id: {}'.format(cov, bin_id))
        return gc, sum_contig_len, cov

    def _generate_contigs(self, file_name, file_directory, assembly_ref):
        """
        _generate_contigs: generate contigs from assembly object

        file_name: file name of fasta file
        file_directory: fasta file directory
        assembly_ref: associated assembly object reference
        """

        log('start generating contig objects for file: {}'.format(file_name))

        assembly = self.dfu.get_objects({'object_refs': [assembly_ref]})['data'][0]
        assembly_contigs = assembly.get('data').get('contigs')

        contigs = {}
        for record in SeqIO.parse(os.path.join(file_directory, file_name), "fasta"):

            contig_id = record.id
            contig = assembly_contigs.get(contig_id)

            if contig:
                # using assembly object data
                contig_gc = contig.get('gc_content')
                sequence_length = contig.get('length')
            else:
                log('cannot find contig [{}] from assembly.'.format(contig_id))
                log('computing contig info')

                sequence = str(record.seq).upper()
                sequence_length = len(sequence)

                contig_gc_len = 0
                contig_gc_len += sequence.count('G')
                contig_gc_len += sequence.count('C')

                contig_gc = round(float(contig_gc_len) / float(sequence_length), 5)

            contig = {
                'gc': contig_gc,
                'len': sequence_length
            }
            contigs[contig_id] = contig

        log('complete generating contig objects for file: {}'.format(file_name))

        return contigs

    def _generate_contig_bin(self, bin_id, file_directory, assembly_ref):
        """
        _generate_contig_bin: gerneate ContigBin structure
        """
        log('start generating BinnedContig info for bin: {}'.format(bin_id))

        # generate ContigBin summery info
        gc, sum_contig_len, cov = self._generate_contig_bin_summary(bin_id, file_directory)

        # generate Contig info
        contigs = self._generate_contigs(bin_id, file_directory, assembly_ref)

        contig_bin = {
            'bid': bin_id,
            'contigs': contigs,
            'n_contigs': len(contigs),
            'gc': gc,
            'sum_contig_len': sum_contig_len,
            'cov': cov
        }

        log('complete generating BinnedContig info for bin: {}'.format(bin_id))

        return contig_bin

    def _get_contig_file(self, assembly_ref):
        """
        _get_contig_file: get contif file from GenomeAssembly object
        """

        log('retrieving contig file from assembly: {}'.format(assembly_ref))
        contig_file = self.au.get_assembly_as_fasta({'ref': assembly_ref}).get('path')

        sys.stdout.flush()
        contig_file = self.dfu.unpack_file({'file_path': contig_file})['file_path']

        log('saved contig file to: {}'.format(contig_file))

        return contig_file

    def _get_contig_string(self, contig_id, assembly_contig_file, parsed_assembly):
        """
        _get_contig_string: find and return contig string from assembly contig file
        """

        # parsed_assembly = SeqIO.to_dict(SeqIO.parse(assembly_contig_file, "fasta"))

        contig_record = parsed_assembly.get(contig_id)

        if contig_record:
            string_contig = ''
            string_contig += '>{}\n'.format(contig_id)
            string_contig += str(contig_record.seq).upper()
            string_contig += '\n'
        else:
            error_msg = 'Cannot find contig [{}] from file [{}].'.format(contig_id,
                                                                         assembly_contig_file)
            raise ValueError(error_msg)

        return string_contig

    def _pack_file_to_shock(self, result_files):
        """
        _pack_file_to_shock: pack files in result_files list and save in shock
        """

        log('start packing and uploading files:\n{}'.format('\n'.join(result_files)))

        output_directory = os.path.join(self.scratch, 'packed_binned_contig_' + str(uuid.uuid4()))
        self._mkdir_p(output_directory)
        result_file = os.path.join(output_directory,
                                   'packed_binned_contig_' + str(uuid.uuid4()) + '.zip')

        with zipfile.ZipFile(result_file, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zip_file:
            for file in result_files:
                zip_file.write(file, os.path.basename(file))

        shock_id = self.dfu.file_to_shock({'file_path': result_file}).get('shock_id')

        log('saved file to shock: {}'.format(shock_id))

        return shock_id

    def _generate_report(self, report_message, params, created_objects=None):
        """
        generate_report: generate summary report

        """
        log('Generating report')

        uuid_string = str(uuid.uuid4())
        upload_message = 'Job Finished\n\n'
        upload_message += report_message

        log('Report message:\n{}'.format(upload_message))

        report_params = {'message': upload_message,
                         'workspace_name': params.get('workspace_name'),
                         'report_object_name': 'MetagenomeUtils_report_' + uuid_string
                         }
        if created_objects:
            report_params['objects_created'] = created_objects

        kbase_report_client = KBaseReport(self.callback_url)
        output = kbase_report_client.create_extended_report(report_params)

        report_output = {'report_name': output['name'], 'report_ref': output['ref']}

        return report_output

    def _generate_report_message(self, new_binned_contig_ref):
        """
        _generate_report_message: generate a report message for BinnedContig object
        """

        report_message = ''

        binned_contig = self.dfu.get_objects({'object_refs': [new_binned_contig_ref]})['data'][0]
        binned_contig_info = binned_contig.get('info')
        binned_contig_name = binned_contig_info[1]
        report_message += 'Generated BinnedContig: {} [{}]\n'.format(binned_contig_name,
                                                                     new_binned_contig_ref)

        binned_contig_count = 0
        total_bins = binned_contig.get('data').get('bins')
        total_bins_count = len(total_bins)
        bin_ids = []
        for bin in total_bins:
            binned_contig_count += len(bin.get('contigs'))
            bin_ids.append(bin.get('bid'))

        report_message += '--------------------------\nSummary:\n\n'
        report_message += 'Binned contigs: {}\n'.format(binned_contig_count)
        report_message += 'Total size of bins: {}\n'.format(total_bins_count)
        report_message += 'Bin IDs:\n{}\n'.format('\n'.join(bin_ids))

        return report_message

    def _merge_bins(self, new_bin_id, bin_objects_to_merge):
        """
        _merge_bins: merge a list of bins into new_bin_id

        """
        total_contigs = {}
        total_gc_count = 0
        total_sum_contig_len = 0
        total_cov_len = 0

        for bin in bin_objects_to_merge:
            total_contigs.update(bin.get('contigs'))
            sum_contig_len = bin.get('sum_contig_len')
            total_sum_contig_len += sum_contig_len
            total_gc_count += sum_contig_len * bin.get('gc')
            total_cov_len += sum_contig_len * bin.get('cov')

        contig_bin = {
            'bid': new_bin_id,
            'contigs': total_contigs,
            'n_contigs': len(total_contigs),
            'gc': round(float(total_gc_count) / total_sum_contig_len, 5),
            'sum_contig_len': total_sum_contig_len,
            'cov': round(float(total_cov_len) / total_sum_contig_len, 5)
        }

        return contig_bin

    def _save_binned_contig(self, binned_contigs, workspace_name, binned_contig_name):
        """
        _build_binned_contig: save BinnedContig object
        """

        workspace_name = workspace_name
        if isinstance(workspace_name, int) or workspace_name.isdigit():
            workspace_id = workspace_name
        else:
            workspace_id = self.dfu.ws_name_to_id(workspace_name)

        object_type = 'KBaseMetagenomes.BinnedContigs'
        save_object_params = {
            'id': workspace_id,
            'objects': [{'type': object_type,
                         'data': binned_contigs,
                         'name': binned_contig_name}]
        }

        dfu_oi = self.dfu.save_objects(save_object_params)[0]
        new_binned_contig_ref = str(dfu_oi[6]) + '/' + str(dfu_oi[0]) + '/' + str(dfu_oi[4])

        return new_binned_contig_ref

    def _check_bin_merges(self, bin_merges):
        """
        _check_bin_merges: checking bin_merges
        """
        bin_id_list = map(lambda item: item.get('bin_to_merge'), bin_merges)
        bin_ids = []
        map(lambda item: map(lambda bin_id: bin_ids.append(bin_id), item), bin_id_list)

        for bin_id in bin_id_list:
            if len(bin_id) <= 1:
                raise ValueError("Please provide at least two bin_ids to merge")
            for id in bin_id:
                if bin_ids.count(id) > 1:
                    raise ValueError("Same bin [{}] appears in muliple merges".format(id))

        new_bin_id_list = map(lambda item: item.get('new_bin_id'), bin_merges)
        for new_bin_id in new_bin_id_list:
            if new_bin_id_list.count(new_bin_id) > 1:
                raise ValueError("Same new Bin ID [{}] appears in muliple merges".format(id))

    def _download_file_from_shock(self, shock_id):
        """
        _download_file_from_shock: download file from shock_id
        """
        log('start downloading file from shock_id: {}'.format(shock_id))

        output_directory = os.path.join(self.scratch, 'binned_contig_file_' + str(uuid.uuid4()))
        self._mkdir_p(output_directory)

        shock_to_file_input = [{'shock_id': shock_id,
                                'file_path': output_directory}]

        shock_to_file_output = self.dfu.shock_to_file_mass(shock_to_file_input)

        file_path = shock_to_file_output[0]["file_path"]
        file_name = shock_to_file_output[0]["node_file_name"]

        return file_path, file_name

    def _process_binned_contig_data(self, binned_contig_data):
        """
        _process_binned_contig_data: construc binned_contig data
        """

        bin_id = binned_contig_data.get('bin_id')[0]

        try:
            sheet_assembly_ref = binned_contig_data.get('assembly_ref')[0]
        except:
            raise ValueError('Unexcepted assembly_ref in Excel sheet: {}'.format(bin_id))
        else:
            if not sheet_assembly_ref:
                raise ValueError('assembly_ref is None in Excel sheet: {}'.format(bin_id))

        cov = binned_contig_data.get('total_coverage')[0]

        contig_ids = binned_contig_data.get('contig_id')
        contig_gcs = binned_contig_data.get('gc')
        contig_lens = binned_contig_data.get('len')
        contig_covs = binned_contig_data.get('contig_coverage')

        contigs = {}

        sum_contig_len = 0
        sum_gc_count = 0
        for index, contig_id in enumerate(contig_ids):
            contig_len = int(contig_lens[index])
            contig_gc = contig_gcs[index]
            # contig_cov = contig_covs[index]
            try:
                contig_cov = float(contig_covs[index])
            except:
                contig_cov = None
            if contig_cov:
                contigs.update({contig_id: {'gc': contig_gc,
                                            'len': contig_len,
                                            'cov': contig_cov}})
            else:
                contigs.update({contig_id: {'gc': contig_gc,
                                            'len': contig_len}})
            sum_gc_count += round(contig_len * contig_gc, 5)
            sum_contig_len += int(contig_len)

        contig_bin = {
            'bid': bin_id,
            'contigs': contigs,
            'n_contigs': len(contigs),
            'gc': round(float(sum_gc_count) / sum_contig_len, 5),
            'sum_contig_len': sum_contig_len,
            'cov': cov
        }

        return sheet_assembly_ref, contig_bin

    def _process_binned_contig_excel(self, file_path):
        """
        _process_binned_contig_excel: fetch and construct BinnedContig info from file
        """
        try:
            workbook = load_workbook(filename=file_path, read_only=True)
        except:
            raise ValueError('Unexpected excel file')

        sheets = workbook.get_sheet_names()

        bins = []
        assembly_ref = None
        total_contig_len = 0

        for sheet in sheets:
            worksheet = workbook[sheet]

            binned_contig_data = {}
            for row in worksheet.rows:
                key_entry = True
                key = ''
                for cell in row:
                    if key_entry:
                        key = cell.value
                        binned_contig_data[key] = []
                        key_entry = False
                    else:
                        binned_contig_data[key].append(cell.value)

            sheet_assembly_ref, contig_bin = self._process_binned_contig_data(binned_contig_data)
            if not assembly_ref:
                assembly_ref = sheet_assembly_ref
            else:
                if assembly_ref != sheet_assembly_ref:
                    raise ValueError('Excel sheets have different assembly_ref')
            bins.append(contig_bin)
            total_contig_len += contig_bin.get('sum_contig_len')

        return assembly_ref, bins, total_contig_len

    def __init__(self, config):
        self.callback_url = config['SDK_CALLBACK_URL']
        self.scratch = config['scratch']
        self.shock_url = config['shock-url']
        self.dfu = DataFileUtil(self.callback_url)
        self.au = AssemblyUtil(self.callback_url)
        self.setapi = SetAPI(self.callback_url)
        self.wss = workspaceService(config['workspace-url'])

    def file_to_binned_contigs(self, params):
        """
        file_to_binned_contigs: Generating BinnedContigs ojbect from files

        input params:
        file_directory: file directory containing compressed/unpacked contig file(s) to
                        build BinnedContig object
        assembly_ref: metagenome assembly object reference
        binned_contig_name: BinnedContig object name
        workspace_name: the name/id of the workspace it gets saved to

        return params:
        binned_contig_obj_ref: generated result BinnedContig object reference
        """

        log('--->\nrunning MetagenomeFileUtils.file_to_binned_contigs\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        self._validate_file_to_binned_contigs_params(params)

        file_directory = params.get('file_directory')
        assembly_ref = params.get('assembly_ref')

        log('starting generating BinnedContig object')
        bin_ids = self._get_bin_ids(file_directory)

        bins = []
        for bin_id in bin_ids:
            contig_bin = self._generate_contig_bin(bin_id, file_directory, assembly_ref)
            bins.append(contig_bin)
        log('finished generating BinnedContig object')

        total_contig_len = self._get_total_contig_len(file_directory)

        binned_contigs = {
            'assembly_ref': assembly_ref,
            'bins': bins,
            'total_contig_len': total_contig_len
        }

        binned_contig_obj_ref = self._save_binned_contig(binned_contigs,
                                                         params.get('workspace_name'),
                                                         params.get('binned_contig_name'))

        returnVal = {'binned_contig_obj_ref': binned_contig_obj_ref}
        log('successfully saved BinnedContig object')

        return returnVal

    def binned_contigs_to_file(self, params):
        """
        binned_contigs_to_file: Convert BinnedContig object to fasta files and pack them to shock

        input params:
        input_ref: BinnedContig object reference

        optional params:
        save_to_shock: saving result bin files to shock. default to True
        bin_id_list: only extract bin_id_list

        return params:
        shock_id: saved packed file shock id
        bin_file_directory: directory that contains all bin files
        """

        log('--->\nrunning MetagenomeFileUtils.binned_contigs_to_file\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        self._validate_binned_contigs_to_file_params(params)

        binned_contig_object = self.dfu.get_objects({'object_refs':
                                                    [params.get('input_ref')]})['data'][0]

        assembly_ref = binned_contig_object.get('data').get('assembly_ref')
        assembly_contig_file = self._get_contig_file(assembly_ref)
        log('parsing assembly file [{}] to dictionary'.format(assembly_contig_file))
        parsed_assembly = SeqIO.to_dict(SeqIO.parse(assembly_contig_file, "fasta"))

        bins = binned_contig_object.get('data').get('bins')

        result_directory = os.path.join(self.scratch, 'binned_contig_files_' + str(uuid.uuid4()))
        self._mkdir_p(result_directory)

        result_files = []
        bin_id_list = params.get('bin_id_list')
        for bin in bins:
            bin_id = bin.get('bid')
            if bin_id_list:
                if bin_id in bin_id_list:
                    log('processing bin: {}'.format(bin_id))
                    with open(os.path.join(result_directory, bin_id), 'w') as file:
                        contigs = bin.get('contigs')
                        for contig_id in contigs.keys():
                            contig_string = self._get_contig_string(contig_id,
                                                                    assembly_contig_file,
                                                                    parsed_assembly)
                            file.write(contig_string)
                    result_files.append(os.path.join(result_directory, bin_id))
                    log('saved contig file to: {}'.format(result_files[-1]))
            else:
                log('processing bin: {}'.format(bin_id))
                with open(os.path.join(result_directory, bin_id), 'w') as file:
                    contigs = bin.get('contigs')
                    for contig_id in contigs.keys():
                        contig_string = self._get_contig_string(contig_id,
                                                                assembly_contig_file,
                                                                parsed_assembly)
                        file.write(contig_string)
                result_files.append(os.path.join(result_directory, bin_id))
                log('saved contig file to: {}'.format(result_files[-1]))

        if params.get('save_to_shock') or params.get('save_to_shock') is None:
            shock_id = self._pack_file_to_shock(result_files)
        else:
            shock_id = None

        returnVal = {'shock_id': shock_id, 'bin_file_directory': result_directory}

        return returnVal

    def export_binned_contigs_as_excel(self, params):
        """
        export_binned_contigs_as_excel: Convert BinnedContig object to an excel file and pack it
                                        to shock

        input params:
        input_ref: BinnedContig object reference

        optional params:
        save_to_shock: saving result bin files to shock. default to True

        return params:
        shock_id: saved packed file shock id
        bin_file_directory: directory that contains all bin files
        """

        log('--->\nrunning MetagenomeFileUtils.export_binned_contigs_as_excel\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        self._validate_binned_contigs_to_file_params(params)

        binned_contig = self.dfu.get_objects({'object_refs':
                                             [params.get('input_ref')]})['data'][0]
        binned_contig_name = binned_contig.get('info')[1]
        binned_contig_data = binned_contig.get('data')

        assembly_ref = binned_contig_data.get('assembly_ref')
        bins = binned_contig_data.get('bins')

        result_directory = os.path.join(self.scratch, 'binned_contig_excel_' + str(uuid.uuid4()))
        self._mkdir_p(result_directory)
        output_excel = os.path.join(result_directory, binned_contig_name + '.xlsx')
        workbook = xlsxwriter.Workbook(output_excel)
        bold = workbook.add_format({'bold': True})

        for bin in bins:
            bin_id = bin.get('bid')
            total_coverage = bin.get('cov', 'NULL')

            log('processing bin: {}'.format(bin_id))
            worksheet = workbook.add_worksheet(bin_id)

            worksheet.write(0, 0, 'bin_id', bold)
            worksheet.write(0, 1, bin_id)

            worksheet.write(1, 0, 'total_coverage', bold)
            worksheet.write(1, 1, total_coverage)

            worksheet.write(2, 0, 'assembly_ref', bold)
            worksheet.write(2, 1, assembly_ref)

            contigs = bin.get('contigs')
            worksheet.write(4, 0, 'contigs', bold)
            worksheet.write(5, 0, 'contig_id', bold)
            worksheet.write(6, 0, 'gc', bold)
            worksheet.write(7, 0, 'len', bold)
            worksheet.write(8, 0, 'contig_coverage', bold)

            i = 1
            for contig_id, contig_value in contigs.items():
                worksheet.write(5, i, contig_id)
                worksheet.write(6, i, contig_value.get('gc'))
                worksheet.write(7, i, contig_value.get('len'))
                worksheet.write(8, i, contig_value.get('cov', 'NULL'))
                i += 1

        workbook.close()

        if params.get('save_to_shock') or params.get('save_to_shock') is None:
            shock_id = self.dfu.file_to_shock({'file_path': output_excel}).get('shock_id')
        else:
            shock_id = None

        returnVal = {'shock_id': shock_id, 'bin_file_directory': result_directory}

        return returnVal

    def import_excel_as_binned_contigs(self, params):
        """
        import_excel_as_binned_contigs: Import an excel file as BinnedContigs

        required params:
        shock_id: Excel file stored in shock
        workspace_name: the name of the workspace object gets saved to

        optional params:
        binned_contigs_name: saved BinnedContig name.
                             Auto append timestamp from excel if not given.
        """

        log('--->\nrunning MetagenomeFileUtils.import_excel_as_binned_contigs\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        self._validate_import_excel_as_binned_contigs_params(params)

        bc_file_path,  bc_file_name = self._download_file_from_shock(params.get('shock_id'))

        binned_contigs_name = params.get('binned_contigs_name')
        if not binned_contigs_name:
            time_stamp = datetime.datetime.fromtimestamp(time.time()).strftime('%Y%m%d_%H%M%S')
            binned_contigs_name = os.path.splitext(bc_file_name)[0] + '_' + time_stamp

        assembly_ref, bins, total_contig_len = self._process_binned_contig_excel(bc_file_path)

        binned_contigs = {
            'assembly_ref': assembly_ref,
            'bins': bins,
            'total_contig_len': total_contig_len
        }

        binned_contigs_ref = self._save_binned_contig(binned_contigs,
                                                      params.get('workspace_name'),
                                                      binned_contigs_name)

        created_objects = []
        created_objects.append({"ref": binned_contigs_ref,
                                "description": "BinnedContigs from Excel"})

        returnVal = {'binned_contigs_ref': binned_contigs_ref}

        report_message = self._generate_report_message(binned_contigs_ref)
        reportVal = self._generate_report(report_message, params, created_objects)
        returnVal.update(reportVal)

        return returnVal

    def _get_object_name_from_ref(self, obj_ref):
        """given the object reference, return the object_name as a string"""
        return(self.wss.get_object_info_new({"objects": [{'ref': obj_ref}]})[0][1])

    def extract_binned_contigs_as_assembly(self, params):
        """
        extract_binned_contigs_as_assembly: extract one/multiple Bins from BinnedContigs as
                                            Assembly

        input params:
        binned_contig_obj_ref: BinnedContig object reference
        extracted_assemblies: a string, a comma-separated list of bin_ids to be extracted
        workspace_name: the name of the workspace it gets saved to

        return params:
        assembly_ref_list: a list of generated result Assembly object reference
        report_name: report name generated by KBaseReport
        report_ref: report reference generated by KBaseReport
        """

        log('--->\nrunning MetagenomeFileUtils.extract_binned_contigs_as_assembly\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        self._validate_extract_binned_contigs_as_assembly_params(params)

        # convert comma-separated list of bins into a list of individual ids (the python
        # comprehension construction deals with the fact that split(',') returns a list of
        # length one, [''], for an empty string input

        extracted_assemblies = [x for x in params.get('extracted_assemblies').split(',') if x]

        binned_contig_obj_ref = params.get('binned_contig_obj_ref')
        contigs_to_file_ret = self.binned_contigs_to_file({'input_ref': binned_contig_obj_ref,
                                                           'save_to_shock': False,
                                                           'bin_id_list': extracted_assemblies})

        bin_file_directory = contigs_to_file_ret.get('bin_file_directory')
        # bin_files will be either a list of the bin contig files corresponding to the
        # target bin ids, or a list of all bin contig files if extracted_assemblies is empty
        bin_files = os.listdir(bin_file_directory)

        # if extracted_assemblies is empty list, create a full one here
        if not extracted_assemblies:
            extracted_assemblies = bin_files
            log("extracted_assemblies was empty, is now " + pformat(extracted_assemblies))

        generated_assembly_ref_list = []
        assembly_suffix = params.get('assembly_suffix').strip()
        for bin_id in extracted_assemblies:
            if bin_id not in map(os.path.basename, bin_files):
                error_msg = 'bin_id [{}] cannot be found in BinnedContig '.format(bin_id)
                error_msg += '[{}]'.format(binned_contig_obj_ref)
                raise ValueError(error_msg)
            else:
                output_assembly_name = bin_id + assembly_suffix
                log('saving assembly: {}'.format(output_assembly_name))
                for bin_file in bin_files:
                    if os.path.basename(bin_file) == bin_id:
                        log('starting generating assembly from {}'.format(bin_id))
                        assembly_params = {
                            'file': {'path': os.path.join(bin_file_directory, bin_file)},
                            'workspace_name': params.get('workspace_name'),
                            'assembly_name': output_assembly_name
                        }
                        assembly_ref = self.au.save_assembly_from_fasta(assembly_params)
                        log('finished generating assembly from {}'.format(bin_id))
                        generated_assembly_ref_list.append(assembly_ref)
        setret = None
        if (len(generated_assembly_ref_list) > 1):
            binned_contig_object_name = self._get_object_name_from_ref(binned_contig_obj_ref)
            assembly_set_name = params.get('assembly_set_name')
            if not assembly_set_name:
                assembly_set_name = "extracted_bins.AssemblySet"
            log("saving assembly set {0}".format(assembly_set_name))
            setret = self.setapi.save_assembly_set_v1({
                'workspace': params.get('workspace_name'),
                'output_object_name': assembly_set_name,
                'data': {
                          'description': 'binned assemblies from {0}'
                                         . format(binned_contig_object_name),
                          'items':  [{'ref': r} for r in generated_assembly_ref_list]
                        }
               })
            log("save assembly set_ref is {0}".format(setret.get('set_ref')))

        report_message = 'Generated Assembly Reference: {}'.format(
            ', '.join(generated_assembly_ref_list))

        # Request from Dylan to add "objects_created" li st to report object
        if setret:
            report_message = report_message+'\nGenerated Assembly Set: {}'.format(
                                                                              setret.get('set_ref'))

        created_objects = []
        if setret:    # if assembly set created put that first
            created_objects.append({"ref": setret.get('set_ref'),
                                    "description": "Assembly set of extracted assemblies"
                                    })
        for assembly_ref in generated_assembly_ref_list:
            created_objects.append({"ref": assembly_ref,
                                    "description": "Assembly object of extracted contigs"
                                    })

        reportVal = self._generate_report(report_message, params, created_objects)

        returnVal = {'assembly_ref_list': generated_assembly_ref_list}
        returnVal.update(reportVal)

        if setret:
            returnVal.update({'assembly_set_ref': setret.get('set_ref')})

        return returnVal

    def remove_bins_from_binned_contig(self, params):
        """
        remove_bins_from_binned_contig: remove a list of bins from BinnedContig object

        input params:
        old_binned_contig_ref: Original BinnedContig object reference
        bins_to_remove: a list of bin ids to be removed
        output_binned_contig_name: Name for the output BinnedContigs object
        workspace_name: the name of the workspace new object gets saved to

        return params:
        new_binned_contig_ref: newly created BinnedContig object referece
        """

        log('--->\nrunning MetagenomeFileUtils.remove_bins_from_binned_contig\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        self._validate_remove_bins_from_binned_contig_params(params)

        binned_contig_object = self.dfu.get_objects({'object_refs':
                                                     [params.get('old_binned_contig_ref')]}
                                                    )['data'][0]

        assembly_ref = binned_contig_object.get('data').get('assembly_ref')
        total_contig_len = int(binned_contig_object.get('data').get('total_contig_len'))

        old_bins = binned_contig_object.get('data').get('bins')
        bins_to_remove = params.get('bins_to_remove')

        for bin in list(old_bins):
            bin_id = bin.get('bid')
            if bin_id in bins_to_remove:
                log('removing bin_id: {}'.format(bin_id))
                old_bins.remove(bin)
                total_contig_len -= int(bin.get('sum_contig_len'))
                log('removed bin_id: {} from BinnedContig object'.format(bin_id))

        binned_contigs = {
            'assembly_ref': assembly_ref,
            'bins': old_bins,
            'total_contig_len': total_contig_len
        }

        new_binned_contig_ref = self._save_binned_contig(binned_contigs,
                                                         params.get('workspace_name'),
                                                         params.get('output_binned_contig_name'))

        returnVal = {'new_binned_contig_ref': new_binned_contig_ref}
        log('successfully saved BinnedContig object')

        return returnVal

    def merge_bins_from_binned_contig(self, params):
        """
        merge_bins_from_binned_contig: merge a list of bins from BinnedContig object

        input params:
        old_binned_contig_ref: Original BinnedContig object reference
        bin_merges: a list of bin merges dicts
            new_bin_id: newly created bin id
            bin_to_merge: list of bins to merge
        output_binned_contig_name: Name for the output BinnedContigs object
        workspace_name: the name of the workspace new object gets saved to

        return params:
        new_binned_contig_ref: newly created BinnedContig object referece
        """

        log('--->\nrunning MetagenomeFileUtils.merge_bins_from_binned_contig\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        self._validate_merge_bins_from_binned_contig_params(params)

        bin_merges = params.get('bin_merges')
        self._check_bin_merges(bin_merges)

        binned_contig_object = self.dfu.get_objects({'object_refs':
                                                     [params.get('old_binned_contig_ref')]}
                                                    )['data'][0]

        assembly_ref = binned_contig_object.get('data').get('assembly_ref')
        total_contig_len = int(binned_contig_object.get('data').get('total_contig_len'))

        bins = binned_contig_object.get('data').get('bins')
        old_bin_ids = map(lambda item: item.get('bid'), bins)

        for bin_merge in bin_merges:
            new_bin_id = bin_merge.get('new_bin_id')
            bin_id_to_merge = bin_merge.get('bin_to_merge')
            if set(bin_id_to_merge) <= set(old_bin_ids):
                bin_objects_to_merge = []
                for bin in list(bins):
                    bin_id = bin.get('bid')
                    if bin_id in bin_id_to_merge:
                        bin_objects_to_merge.append(bin)
                        log('removing bin_id: {}'.format(bin_id))
                        bins.remove(bin)
                        total_contig_len -= int(bin.get('sum_contig_len'))
                        log('removed bin_id: {} from BinnedContig object'.format(bin_id))
                new_bin = self._merge_bins(new_bin_id, bin_objects_to_merge)
                log('appending bin_id: {}'.format(new_bin_id))
                bins.append(new_bin)
                total_contig_len += int(new_bin.get('sum_contig_len'))
                log('appended bin_id: {} to BinnedContig object'.format(new_bin_id))
            else:
                bad_bin_ids = list(set(bin_id_to_merge) - set(old_bin_ids))
                error_msg = 'bin_id: [{}] '.format(', '.join(bad_bin_ids))
                error_msg += 'is not listed in BinnedContig object'
                raise ValueError(error_msg)

        binned_contigs = {
            'assembly_ref': assembly_ref,
            'bins': bins,
            'total_contig_len': total_contig_len
        }

        new_binned_contig_ref = self._save_binned_contig(binned_contigs,
                                                         params.get('workspace_name'),
                                                         params.get('output_binned_contig_name'))

        returnVal = {'new_binned_contig_ref': new_binned_contig_ref}
        log('successfully saved BinnedContig object')

        return returnVal

    def edit_bins_from_binned_contig(self, params):
        """
        edit_bins_from_binned_contig: merge/remove a list of bins from BinnedContig object
                                    a wrapper method of:
                                    merge_bins_from_binned_contig
                                    remove_bins_from_binned_contig


        input params:
        old_binned_contig_ref: Original BinnedContig object reference
        bins_to_remove: a list of bin ids to be removed
        bin_merges: a list of bin merges dicts
            new_bin_id: newly created bin id
            bin_to_merge: list of bins to merge
        output_binned_contig_name: Name for the output BinnedContigs object
        workspace_name: the name of the workspace new object gets saved to

        return params:
        new_binned_contig_ref: newly created BinnedContig object referece
        report_name: report name generated by KBaseReport
        report_ref: report reference generated by KBaseReport
        """

        log('--->\nrunning MetagenomeFileUtils.edit_bins_from_binned_contig\n' +
            'params:\n{}'.format(json.dumps(params, indent=1)))

        input_params = params.copy()
        if params.get('bins_to_remove'):
            bins_to_remove = input_params.get('bins_to_remove')
            if isinstance(bins_to_remove, string_types):
                input_params['bins_to_remove'] = bins_to_remove.split(',')
            new_binned_contig_ref = self.remove_bins_from_binned_contig(input_params).get(
                'new_binned_contig_ref')
            input_params['old_binned_contig_ref'] = new_binned_contig_ref

        if params.get('bin_merges'):
            new_binned_contig_ref = self.merge_bins_from_binned_contig(input_params).get(
                'new_binned_contig_ref')

        returnVal = {'new_binned_contig_ref': new_binned_contig_ref}

        report_message = self._generate_report_message(new_binned_contig_ref)
        reportVal = self._generate_report(report_message, params)
        returnVal.update(reportVal)

        return returnVal
